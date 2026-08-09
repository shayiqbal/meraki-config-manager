"""Import CSV, JSON, text JSON, and XLSX rule files."""
from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from pydantic import ValidationError

from rules.models import RuleCategory, RuleSet, RuleSource, VpnExclusionRule


class ImportValidationError(ValueError):
    """Raised with row/rule-specific import diagnostics."""


def _category(value: str) -> RuleCategory:
    key = value.strip().lower().replace("_", "")
    mapping = {
        "custom": RuleCategory.CUSTOM,
        "customlayer3": RuleCategory.CUSTOM,
        "layer3": RuleCategory.CUSTOM,
        "dns": RuleCategory.CUSTOM,
        "majorapplication": RuleCategory.MAJOR_APPLICATION,
        "majorapplications": RuleCategory.MAJOR_APPLICATION,
        "application": RuleCategory.APPLICATION,
        "applications": RuleCategory.APPLICATION,
    }
    try:
        return mapping[key]
    except KeyError as exc:
        raise ImportValidationError(f"unsupported ruleType {value!r}") from exc


def _source_from(data: dict[str, Any]) -> RuleSource | None:
    source = data.get("source")
    if isinstance(source, dict):
        return RuleSource.model_validate(source)
    cidr = data.get("sourceCidr") or data.get("source_cidr")
    port = data.get("sourcePort") or data.get("source_port")
    vlan = data.get("sourceVlan") or data.get("source_vlan") or data.get("source")
    if isinstance(vlan, str) and vlan.strip().lower() in {"", "any"}:
        vlan = None
    return RuleSource(cidr=cidr or None, port=port or None, vlanId=vlan or None) if any(
        (cidr, port, vlan)
    ) else None


def _rule(data: dict[str, Any], index: int, origin: str) -> VpnExclusionRule:
    category = _category(
        str(data.get("category") or data.get("ruleType") or data.get("type") or "")
    )
    protocol = str(data.get("protocol") or "any")
    if str(data.get("ruleType", "")).lower() == "dns":
        protocol = "dns"
    try:
        return VpnExclusionRule(
            category=category,
            order=int(data.get("order") or data.get("ruleOrder") or index),
            protocol=protocol,
            destination=data.get("destination"),
            port=str(data.get("port") or "any"),
            application_id=data.get("application_id") or data.get("applicationId") or data.get("id"),
            name=data.get("name"),
            source=_source_from(data),
            action=str(data.get("action") or "upsert"),
            origin=f"{origin} row/rule {index}",
        )
    except (ValidationError, ValueError, TypeError) as exc:
        raise ImportValidationError(f"{origin} row/rule {index}: {exc}") from exc


def _from_rows(rows: Iterable[dict[str, Any]], origin: str) -> RuleSet:
    rules: list[VpnExclusionRule] = []
    errors: list[str] = []
    for index, row in enumerate(rows, 2):
        if not any(value not in (None, "") for value in row.values()):
            continue
        try:
            rules.append(_rule(row, index, origin))
        except ImportValidationError as exc:
            errors.append(str(exc))
    if errors:
        raise ImportValidationError("\n".join(errors))
    return RuleSet(rules=rules)


def parse_csv(path: Path) -> RuleSet:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return _from_rows(csv.DictReader(handle), path.name)


def parse_xlsx(path: Path) -> RuleSet:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return RuleSet()
    headers = [str(item or "").strip() for item in rows[0]]
    return _from_rows(
        (dict(zip(headers, row, strict=False)) for row in rows[1:]), path.name
    )


def parse_json_data(data: Any, origin: str) -> RuleSet:
    metadata: dict[str, Any] = {}
    mode = "merge"
    if isinstance(data, list):
        return RuleSet(
            rules=[_rule(item, index, origin) for index, item in enumerate(data, 1)]
        )
    if not isinstance(data, dict):
        raise ImportValidationError(f"{origin}: JSON root must be an object or array")
    metadata = data.get("exportMetadata", data.get("metadata", {}))
    mode = data.get("mode", "merge")
    container = data.get("localInternetBreakout", data)
    if "rules" in container:
        rules = [
            _rule(item, index, origin)
            for index, item in enumerate(container["rules"], 1)
        ]
    else:
        rules = []
        for category in RuleCategory:
            for item in container.get(category.value, []):
                converted = dict(item)
                converted["category"] = category.value
                rules.append(_rule(converted, len(rules) + 1, origin))
    return RuleSet(rules=rules, mode=mode, metadata=metadata)


def parse_json(path: Path) -> RuleSet:
    try:
        return parse_json_data(json.loads(path.read_text(encoding="utf-8-sig")), path.name)
    except json.JSONDecodeError as exc:
        raise ImportValidationError(
            f"{path.name}: invalid JSON at line {exc.lineno}, column {exc.colno}: {exc.msg}"
        ) from exc


def parse_file(path: str | Path) -> RuleSet:
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return parse_csv(path)
    if suffix in {".json", ".txt"}:
        return parse_json(path)
    if suffix == ".xlsx":
        return parse_xlsx(path)
    raise ImportValidationError("supported formats are CSV, JSON, TXT (JSON), and XLSX")

