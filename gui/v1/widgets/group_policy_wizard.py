"""Group Policy wizard — 5-step guided workflow for multi-network policy copying."""
from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QStackedWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from services.group_policy_service import (
    GroupPolicyCopyService,
    NetworkPolicyPreview,
    PolicyCopyResult,
    _summarise,
)

if TYPE_CHECKING:
    from gui.v1.main_window_v1 import MainWindowV1

_STEP_LABELS = [
    "Select Source",
    "Pick Policies",
    "Pick Destinations",
    "Preview",
    "Results",
]


def _step_header(step: int, total: int, title: str, subtitle: str) -> QWidget:
    w = QWidget()
    lay = QVBoxLayout(w)
    lay.setContentsMargins(0, 0, 0, 8)
    lay.setSpacing(2)
    step_note = QLabel(f"Step {step} of {total}  —  {title}")
    step_note.setObjectName("section-title")
    sub = QLabel(subtitle)
    sub.setObjectName("section-sub")
    sub.setWordWrap(True)
    lay.addWidget(step_note)
    lay.addWidget(sub)
    return w


class GroupPolicyWizardPanel(QWidget):
    """5-step wizard: source → policies → destinations → preview → results."""

    def __init__(self, app: "MainWindowV1", parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._app = app
        self._networks: list[dict[str, Any]] = []
        self._source_policies: list[dict[str, Any]] = []
        self._selected_policies: list[dict[str, Any]] = []
        self._previews: list[NetworkPolicyPreview] = []
        self._results: list[PolicyCopyResult] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(28, 20, 28, 20)
        root.setSpacing(0)

        # ── Step indicator ─────────────────────────────────────────────────────
        self._indicator = self._build_indicator()
        root.addWidget(self._indicator)

        # ── Stacked pages ──────────────────────────────────────────────────────
        self._stack = QStackedWidget()
        self._page_source  = self._build_source()
        self._page_policies = self._build_policies()
        self._page_dests   = self._build_dests()
        self._page_preview = self._build_preview()
        self._page_results = self._build_results()
        for page in (
            self._page_source, self._page_policies,
            self._page_dests, self._page_preview, self._page_results,
        ):
            self._stack.addWidget(page)
        root.addWidget(self._stack, 1)

        # ── Navigation buttons ─────────────────────────────────────────────────
        nav = QHBoxLayout()
        nav.setContentsMargins(0, 12, 0, 0)
        self._back_btn = QPushButton("← Back")
        self._back_btn.clicked.connect(self._go_back)
        self._next_btn = QPushButton("Next →")
        self._next_btn.setObjectName("primary")
        self._next_btn.clicked.connect(self._go_next)
        self._restart_btn = QPushButton("Start Over")
        self._restart_btn.clicked.connect(lambda: self._goto(0))
        self._restart_btn.setVisible(False)
        nav.addWidget(self._back_btn)
        nav.addStretch()
        nav.addWidget(self._restart_btn)
        nav.addWidget(self._next_btn)
        root.addLayout(nav)

        self._goto(0)

    # ── Page builders ──────────────────────────────────────────────────────────

    def _build_indicator(self) -> QWidget:
        w = QWidget()
        lay = QHBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 16)
        lay.setSpacing(0)
        self._step_labels: list[QLabel] = []
        for i, label in enumerate(_STEP_LABELS):
            num = QLabel(str(i + 1))
            num.setObjectName("step-inactive")
            num.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self._step_labels.append(num)
            lay.addWidget(num)
            text = QLabel(f"  {label}  ")
            text.setObjectName("section-sub")
            lay.addWidget(text)
            if i < len(_STEP_LABELS) - 1:
                divider = QLabel("──")
                divider.setObjectName("section-sub")
                lay.addWidget(divider)
        lay.addStretch()
        return w

    def _build_source(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(10)
        lay.addWidget(_step_header(
            1, 5, "Select Source Network",
            "Choose the network whose Group Policies you want to compare and copy.",
        ))
        self._source_search = QLineEdit()
        self._source_search.setObjectName("search")
        self._source_search.setPlaceholderText("🔍  Search networks…")
        self._source_search.textChanged.connect(self._filter_source_list)
        lay.addWidget(self._source_search)
        self._source_list = QListWidget()
        self._source_list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._source_list.itemSelectionChanged.connect(self._on_source_selected)
        lay.addWidget(self._source_list)
        self._source_status = QLabel("")
        self._source_status.setObjectName("section-sub")
        lay.addWidget(self._source_status)
        return w

    def _build_policies(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(10)
        lay.addWidget(_step_header(
            2, 5, "Select Policies to Copy",
            "Check the Group Policies you want to copy to other networks.",
        ))
        sel_row = QHBoxLayout()
        sel_all = QPushButton("Select All")
        sel_all.clicked.connect(lambda: self._check_all_policies(True))
        clr = QPushButton("Clear")
        clr.clicked.connect(lambda: self._check_all_policies(False))
        sel_row.addWidget(sel_all)
        sel_row.addWidget(clr)
        sel_row.addStretch()
        lay.addLayout(sel_row)

        self._policies_table = QTableWidget(0, 5)
        self._policies_table.setHorizontalHeaderLabels(
            ["✓", "Policy Name", "Bandwidth", "Scheduling", "VLAN / Firewall"]
        )
        self._policies_table.setAlternatingRowColors(True)
        self._policies_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch
        )
        self._policies_table.horizontalHeader().setSectionResizeMode(
            4, QHeaderView.ResizeMode.Stretch
        )
        self._policies_table.verticalHeader().setVisible(False)
        self._policies_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        lay.addWidget(self._policies_table)

        self._policies_status = QLabel("")
        self._policies_status.setObjectName("section-sub")
        lay.addWidget(self._policies_status)
        return w

    def _build_dests(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(10)
        lay.addWidget(_step_header(
            3, 5, "Select Destination Networks",
            "Choose one or more networks to receive the selected policies. "
            "The source network is excluded automatically.",
        ))
        self._dests_search = QLineEdit()
        self._dests_search.setObjectName("search")
        self._dests_search.setPlaceholderText("🔍  Search networks…")
        self._dests_search.textChanged.connect(self._filter_dests_list)
        sel_row = QHBoxLayout()
        sel_all = QPushButton("Select All")
        sel_all.clicked.connect(lambda: self._check_all_dests(True))
        clr = QPushButton("Clear")
        clr.clicked.connect(lambda: self._check_all_dests(False))
        sel_row.addWidget(sel_all)
        sel_row.addWidget(clr)
        sel_row.addStretch()
        lay.addWidget(self._dests_search)
        lay.addLayout(sel_row)
        self._dests_list = QListWidget()
        self._dests_list.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        lay.addWidget(self._dests_list)
        return w

    def _build_preview(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(10)
        lay.addWidget(_step_header(
            4, 5, "Preview Changes",
            "Review every policy that will be added or skipped per network. "
            "Type CONFIRM below to enable deployment.",
        ))
        self._preview_text = QTextEdit()
        self._preview_text.setReadOnly(True)
        lay.addWidget(self._preview_text)

        # CONFIRM gate
        confirm_row = QHBoxLayout()
        confirm_lbl = QLabel("Type  CONFIRM  to enable deployment:")
        confirm_lbl.setObjectName("label")
        self._confirm_edit = QLineEdit()
        self._confirm_edit.setPlaceholderText("CONFIRM")
        self._confirm_edit.setMaximumWidth(180)
        self._confirm_edit.textChanged.connect(self._gate_apply)
        confirm_row.addWidget(confirm_lbl)
        confirm_row.addWidget(self._confirm_edit)
        confirm_row.addStretch()
        lay.addLayout(confirm_row)

        self._apply_btn = QPushButton("✓ Deploy to All Destinations")
        self._apply_btn.setObjectName("success")
        self._apply_btn.setEnabled(False)
        self._apply_btn.clicked.connect(self._apply)
        self._progress = QProgressBar()
        self._progress.setRange(0, 100)
        self._progress.setVisible(False)
        apply_row = QHBoxLayout()
        apply_row.addStretch()
        apply_row.addWidget(self._progress)
        apply_row.addWidget(self._apply_btn)
        lay.addLayout(apply_row)
        return w

    def _build_results(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(10)
        lay.addWidget(_step_header(
            5, 5, "Results",
            "Deployment complete. Policies are listed as Added or Already Existed per network.",
        ))
        self._results_text = QTextEdit()
        self._results_text.setReadOnly(True)
        lay.addWidget(self._results_text)

        export_row = QHBoxLayout()
        self._export_csv_btn = QPushButton("⬇  Export CSV")
        self._export_csv_btn.setEnabled(False)
        self._export_csv_btn.clicked.connect(self._export_csv)
        self._export_xlsx_btn = QPushButton("⬇  Export Excel")
        self._export_xlsx_btn.setEnabled(False)
        self._export_xlsx_btn.clicked.connect(self._export_excel)
        export_row.addStretch()
        export_row.addWidget(self._export_csv_btn)
        export_row.addWidget(self._export_xlsx_btn)
        lay.addLayout(export_row)
        return w

    # ── Navigation ─────────────────────────────────────────────────────────────

    def _goto(self, index: int) -> None:
        self._stack.setCurrentIndex(index)
        self._back_btn.setVisible(0 < index < 4)
        self._next_btn.setVisible(index < 3)
        self._restart_btn.setVisible(index == 4)
        self._update_indicator(index)
        if index == 0:
            self._populate_source_list()
        elif index == 2:
            self._populate_dests_list()
        elif index == 3:
            self._generate_preview()

    def _go_back(self) -> None:
        self._goto(self._stack.currentIndex() - 1)

    def _go_next(self) -> None:
        idx = self._stack.currentIndex()
        if idx == 0 and not self._validate_source():
            return
        if idx == 1:
            self._collect_selected_policies()
            if not self._selected_policies:
                QMessageBox.warning(self, "No Policies Selected",
                                    "Please check at least one policy to copy.")
                return
        if idx == 2 and not self._get_selected_dests():
            QMessageBox.warning(self, "No Destinations",
                                "Please select at least one destination network.")
            return
        self._goto(idx + 1)

    def _update_indicator(self, active: int) -> None:
        for i, lbl in enumerate(self._step_labels):
            if i < active:
                lbl.setObjectName("step-done")
                lbl.setText("✓")
            elif i == active:
                lbl.setObjectName("step-active")
                lbl.setText(str(i + 1))
            else:
                lbl.setObjectName("step-inactive")
                lbl.setText(str(i + 1))
            lbl.setStyleSheet("")  # force QSS repaint

    # ── Step 1: source list ────────────────────────────────────────────────────

    def refresh(self, networks: list[dict[str, Any]]) -> None:
        self._networks = networks
        if self._stack.currentIndex() == 0:
            self._populate_source_list()

    def _populate_source_list(self) -> None:
        self._source_search.clear()
        self._source_list.clear()
        for net in self._networks:
            item = QListWidgetItem(
                f"{net['name']}  [{', '.join(net.get('productTypes') or [])}]"
            )
            item.setData(Qt.ItemDataRole.UserRole, net)
            self._source_list.addItem(item)

    def _filter_source_list(self, text: str) -> None:
        q = text.lower()
        for i in range(self._source_list.count()):
            item = self._source_list.item(i)
            net = item.data(Qt.ItemDataRole.UserRole)
            item.setHidden(bool(q) and q not in net["name"].lower())

    def _on_source_selected(self) -> None:
        net = self._selected_source()
        if net:
            self._source_status.setText(f"Selected: {net['name']}")
            self._load_source_policies(net)

    def _selected_source(self) -> dict[str, Any] | None:
        items = self._source_list.selectedItems()
        return items[0].data(Qt.ItemDataRole.UserRole) if items else None

    def _validate_source(self) -> bool:
        if not self._selected_source():
            QMessageBox.warning(self, "No Source", "Please select a source network.")
            return False
        return True

    def _load_source_policies(self, network: dict[str, Any]) -> None:
        client = self._app.get_client()
        if not client:
            return
        self._source_status.setText(f"Loading policies for {network['name']}…")
        self._app.run_worker(
            lambda: client.get_group_policies(network["id"]),
            on_result=self._on_source_policies,
            status_msg=f"Loading group policies for {network['name']}…",
        )

    def _on_source_policies(self, policies: list[dict[str, Any]]) -> None:
        self._source_policies = policies
        self._populate_policies_table()
        net = self._selected_source()
        self._source_status.setText(
            f"{len(self._source_policies)} policy(ies) in "
            f"{net['name'] if net else '—'}"
        )

    # ── Step 2: policies table ─────────────────────────────────────────────────

    def _populate_policies_table(self) -> None:
        self._policies_table.setRowCount(0)
        for policy in self._source_policies:
            row = self._policies_table.rowCount()
            self._policies_table.insertRow(row)

            chk = QCheckBox()
            chk.setChecked(True)
            self._policies_table.setCellWidget(row, 0, chk)

            # Name
            self._policies_table.setItem(
                row, 1, QTableWidgetItem(policy.get("name", "—"))
            )

            # Bandwidth
            bw = policy.get("bandwidth", {})
            bw_s = bw.get("settings", "network default")
            if bw_s == "custom":
                limits = bw.get("bandwidthLimits", {})
                up   = limits.get("limitUp")
                down = limits.get("limitDown")
                up_s   = f"↑{up // 1000}M"   if up   else "↑∞"
                down_s = f"↓{down // 1000}M" if down else "↓∞"
                bw_text = f"{up_s} / {down_s}"
            else:
                bw_text = bw_s
            self._policies_table.setItem(row, 2, QTableWidgetItem(bw_text))

            # Scheduling
            sched = policy.get("scheduling", {})
            sched_text = "Enabled" if sched.get("enabled") else "Disabled"
            self._policies_table.setItem(row, 3, QTableWidgetItem(sched_text))

            # VLAN / Firewall summary
            vlan = policy.get("vlanTagging", {})
            vlan_s = vlan.get("settings", "network default")
            vlan_text = f"VLAN {vlan.get('vlanId', '?')}" if vlan_s == "custom" else vlan_s

            fw = policy.get("firewallAndTrafficShaping", {})
            fw_s = fw.get("settings", "network default")
            if fw_s == "custom":
                l3 = len(fw.get("l3FirewallRules", []))
                l7 = len(fw.get("l7FirewallRules", []))
                fw_text = f"FW: {l3}L3 / {l7}L7"
            else:
                fw_text = f"FW: {fw_s}"

            summary_text = f"{vlan_text}  |  {fw_text}"
            self._policies_table.setItem(row, 4, QTableWidgetItem(summary_text))

            # Store the full policy object on the name cell
            self._policies_table.item(row, 1).setData(
                Qt.ItemDataRole.UserRole, policy
            )

    def _check_all_policies(self, checked: bool) -> None:
        for row in range(self._policies_table.rowCount()):
            w = self._policies_table.cellWidget(row, 0)
            if isinstance(w, QCheckBox):
                w.setChecked(checked)

    def _collect_selected_policies(self) -> None:
        self._selected_policies = []
        for row in range(self._policies_table.rowCount()):
            w = self._policies_table.cellWidget(row, 0)
            if isinstance(w, QCheckBox) and w.isChecked():
                policy = self._policies_table.item(row, 1).data(
                    Qt.ItemDataRole.UserRole
                )
                if policy:
                    self._selected_policies.append(policy)
        self._policies_status.setText(
            f"{len(self._selected_policies)} policy(ies) selected"
        )

    # ── Step 3: destination list ───────────────────────────────────────────────

    def _populate_dests_list(self) -> None:
        self._dests_search.clear()
        source = self._selected_source()
        source_id = source["id"] if source else None
        self._dests_list.clear()
        for net in self._networks:
            if net["id"] == source_id:
                continue
            item = QListWidgetItem(net["name"])
            item.setData(Qt.ItemDataRole.UserRole, net)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Unchecked)
            self._dests_list.addItem(item)

    def _filter_dests_list(self, text: str) -> None:
        q = text.lower()
        for i in range(self._dests_list.count()):
            item = self._dests_list.item(i)
            net = item.data(Qt.ItemDataRole.UserRole)
            item.setHidden(bool(q) and q not in net["name"].lower())

    def _check_all_dests(self, checked: bool) -> None:
        state = Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked
        for i in range(self._dests_list.count()):
            item = self._dests_list.item(i)
            if not item.isHidden():
                item.setCheckState(state)

    def _get_selected_dests(self) -> list[dict[str, Any]]:
        return [
            self._dests_list.item(i).data(Qt.ItemDataRole.UserRole)
            for i in range(self._dests_list.count())
            if not self._dests_list.item(i).isHidden()
            and self._dests_list.item(i).checkState() == Qt.CheckState.Checked
        ]

    # ── Step 4: preview ────────────────────────────────────────────────────────

    def _generate_preview(self) -> None:
        dests = self._get_selected_dests()
        policies = self._selected_policies
        if not self._app.get_client():
            return
        self._preview_text.setPlainText("Generating preview…")
        self._apply_btn.setEnabled(False)
        svc = self._app.get_group_policy_service()
        self._app.run_worker(
            lambda: svc.preview(policies, dests),
            on_result=self._on_preview,
            status_msg="Generating group policy preview…",
        )

    def _gate_apply(self) -> None:
        ok = (
            self._confirm_edit.text().strip().upper() == "CONFIRM"
            and bool(self._previews)
        )
        self._apply_btn.setEnabled(ok)

    def _on_preview(self, previews: list[NetworkPolicyPreview]) -> None:
        self._previews = previews
        lines: list[str] = []
        for p in previews:
            lines.append(f"▶  {p.network_name}")
            lines.append(
                f"   Changes: {p.new_count} new policy(ies) will be added  |  "
                f"{p.exists_count} already exist (skipped)  |  "
                f"{p.invalid_count} invalid"
            )
            lines.append("")
            for s in p.policy_statuses:
                badge = {
                    "new":     "[NEW]    ",
                    "exists":  "[EXISTS] ",
                    "invalid": "[INVALID]",
                }[s.status]
                name = s.policy.get("name", "—")
                summary = _summarise(s.policy)
                lines.append(f"   {badge}  {name}")
                lines.append(f"             {summary}")
            lines.append("")

        total_new = sum(p.new_count for p in previews)
        self._preview_text.setPlainText("\n".join(lines))
        self._confirm_edit.clear()
        self._apply_btn.setEnabled(False)
        if total_new == 0:
            self._apply_btn.setText("No New Policies to Apply")
        else:
            self._apply_btn.setText(
                f"✓ Deploy {total_new} Policy(ies) to {len(previews)} Network(s)"
            )

    # ── Step 5: apply ──────────────────────────────────────────────────────────

    def _apply(self) -> None:
        dests = self._get_selected_dests()
        net_names = "\n".join(f"  • {d['name']}" for d in dests)
        answer = QMessageBox.question(
            self,
            "Confirm Copy",
            f"Copy {len(self._selected_policies)} policy(ies) to "
            f"{len(dests)} network(s)?\n\n"
            f"{net_names}\n\n"
            "Policies that already exist on a destination will be skipped automatically.",
        )
        if answer != QMessageBox.StandardButton.Yes:
            return

        self._apply_btn.setEnabled(False)
        self._progress.setVisible(True)
        self._progress.setValue(0)
        policies = self._selected_policies
        svc = self._app.get_group_policy_service()

        def _do() -> list[PolicyCopyResult]:
            return svc.execute(policies, dests)

        def _done(results: list[PolicyCopyResult]) -> None:
            self._progress.setValue(100)
            self._results = results
            self._show_results()
            self._goto(4)
            total_added = sum(r.policies_added for r in results)
            ok_nets = sum(r.success for r in results)
            self._app.log_activity(
                f"Group Policy copy complete: {total_added} policy(ies) added "
                f"across {ok_nets}/{len(results)} network(s).",
                "SUCCESS" if all(r.success for r in results) else "WARNING",
            )

        self._app.run_worker(_do, on_result=_done, status_msg="Copying group policies…")

    def _show_results(self) -> None:
        lines: list[str] = []
        total_added   = sum(r.policies_added  for r in self._results)
        total_skipped = sum(r.policies_skipped for r in self._results)
        ok_nets  = sum(r.success for r in self._results)
        total_nets = len(self._results)

        lines.append(
            f"Deployment complete: {total_added} policy(ies) added, "
            f"{total_skipped} already existed, across {ok_nets}/{total_nets} network(s)"
        )
        lines.append("")

        preview_by_id = {p.network_id: p for p in self._previews}

        for res in self._results:
            icon = "✓" if res.success else "✗"
            lines.append(f"{icon}  {res.network_name}")

            if res.success:
                lines.append(
                    f"   {res.policies_added} policy(ies) added  |  "
                    f"{res.policies_skipped} already existed (skipped)"
                )
                preview = preview_by_id.get(res.network_id)
                if preview:
                    added   = [s for s in preview.policy_statuses if s.status == "new"]
                    existed = [s for s in preview.policy_statuses if s.status == "exists"]
                    invalid = [s for s in preview.policy_statuses if s.status == "invalid"]

                    if added:
                        lines.append("")
                        lines.append("   ADDED:")
                        lines.append("   " + "─" * 72)
                        for s in added:
                            lines.append(
                                f"   + {s.policy.get('name', '—')}"
                            )
                            lines.append(
                                f"     {_summarise(s.policy)}"
                            )

                    if existed:
                        lines.append("")
                        lines.append("   ALREADY EXISTED (skipped):")
                        lines.append("   " + "─" * 72)
                        for s in existed:
                            lines.append(
                                f"   = {s.policy.get('name', '—')}"
                            )

                    if invalid:
                        lines.append("")
                        lines.append("   INVALID (not deployed):")
                        lines.append("   " + "─" * 72)
                        for s in invalid:
                            lines.append(
                                f"   ! {s.policy.get('name', '—')}  — {s.detail}"
                            )
            else:
                lines.append(f"   Error: {res.error}")

            lines.append("")

        self._results_text.setPlainText("\n".join(lines))
        self._export_csv_btn.setEnabled(True)
        self._export_xlsx_btn.setEnabled(True)

    # ── Export helpers ─────────────────────────────────────────────────────────

    def _build_export_rows(self) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        preview_by_id = {p.network_id: p for p in self._previews}
        result_by_id  = {r.network_id: r for r in self._results}

        for net_id, preview in preview_by_id.items():
            res = result_by_id.get(net_id)
            deploy_status = (
                "Success" if (res and res.success)
                else ("Failed" if res else "Unknown")
            )
            for s in preview.policy_statuses:
                policy = s.policy
                bw = policy.get("bandwidth", {})
                bw_s = bw.get("settings", "network default")
                if bw_s == "custom":
                    limits = bw.get("bandwidthLimits", {})
                    up   = limits.get("limitUp")
                    down = limits.get("limitDown")
                    bw_text = (
                        f"↑{up // 1000}M / ↓{down // 1000}M"
                        if (up and down) else bw_s
                    )
                else:
                    bw_text = bw_s
                sched = policy.get("scheduling", {})
                rows.append({
                    "Network":        preview.network_name,
                    "Deploy Status":  deploy_status,
                    "Policy Status":  (
                        "Added" if s.status == "new"
                        else "Already Existed" if s.status == "exists"
                        else "Invalid"
                    ),
                    "Policy Name":    policy.get("name", "—"),
                    "Bandwidth":      bw_text,
                    "Scheduling":     "Enabled" if sched.get("enabled") else "Disabled",
                    "Detail":         s.detail,
                })
        return rows

    def _export_csv(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "Save CSV Report",
            f"group_policies_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            "CSV files (*.csv)",
        )
        if not path:
            return
        rows = self._build_export_rows()
        if not rows:
            return
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
        self._app.log_activity(
            f"Group policy results exported to CSV: {Path(path).name}", "SUCCESS"
        )

    def _export_excel(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel Report",
            f"group_policies_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            "Excel files (*.xlsx)",
        )
        if not path:
            return
        rows = self._build_export_rows()
        if not rows:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(
                self, "Missing Dependency",
                "openpyxl is required for Excel export.\nRun: pip install openpyxl",
            )
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Group Policy Report"

        headers = list(rows[0].keys())
        hdr_fill    = PatternFill("solid", fgColor="1E293B")
        hdr_font    = Font(bold=True, color="F1F5F9", size=11)
        added_fill  = PatternFill("solid", fgColor="DCFCE7")
        exists_fill = PatternFill("solid", fgColor="F1F5F9")
        invalid_fill = PatternFill("solid", fgColor="FEE2E2")

        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in rows:
            ws.append(list(row.values()))
            r = ws.max_row
            status = row.get("Policy Status", "")
            fill = (
                added_fill  if status == "Added"
                else exists_fill  if status == "Already Existed"
                else invalid_fill
            )
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=r, column=col_idx).fill = fill

        for col_idx, header in enumerate(headers, 1):
            max_len = max(
                len(str(header)),
                max((len(str(row.get(header, ""))) for row in rows), default=0),
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

        wb.save(path)
        self._app.log_activity(
            f"Group policy results exported to Excel: {Path(path).name}", "SUCCESS"
        )
